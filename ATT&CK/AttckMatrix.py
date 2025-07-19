#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   AttckMatrix.py
@Time    :   2024/08/03 09:11:03
@Author  :   wwb 
@Version :   1.0
@Contact :   weibo0920@gmail.com
'''

# here put the import lib
import requests
from requests import RequestException
from bs4 import BeautifulSoup
import re
import openpyxl
import json
import time


target = "https://attack.mitre.org"


def request_url(url):
    try:
        resp = requests.get(target + url, timeout=10)
    except RequestException as e:
        print(e)
    bs = BeautifulSoup(resp.content, "html.parser")
    return bs


# 元素去重
def remove_duplicates(lst):
    return list(dict.fromkeys(lst))


# 解析首页获取所有战术
def get_tactics(bs):
    tactics = bs.select(f'td[class="tactic name"]')
    tactics_list = []
    for t in tactics:
        a_tag = t.find('a')
        data = {'href': a_tag.get('href'), 'name': a_tag.text}
        tactics_list.append(json.dumps(data))
    return tactics_list


# 解析战术页面获取技术
def get_technique(bs):
    technique = bs.select(f'tr[class="technique"]')
    result = []
    for i in range(len(technique)):
        a_tag = technique[i].find_all('a')
        data = {'name': a_tag[1].text.strip(), 'href': a_tag[0].get("href")}
        result.append(data)
    return result


# 解析子技术页面获取子技术
def get_sub_technique(bs):
    if bs.find_all("a", class_="subtechnique-table-item"):
        tr_tag = bs.find_all("tbody")[0].find_all("tr")
        result = []
        for t in tr_tag:
            sub_technique = t.find_all("td")[1].text
            result.append(sub_technique.split("\n")[1].strip())
        return result


if __name__ == "__main__":
    bs_tactics = request_url("/versions/v15/")
    tactics_list = remove_duplicates(get_tactics(bs_tactics))
    wb = openpyxl.Workbook()
    sheet = wb.active
    wb.remove(sheet)
    for i, h in enumerate(tactics_list):
        h = json.loads(h)
        ws = wb.create_sheet(h["name"])  # 使用战术名字创建sheet
        bs_technique = request_url(h["href"])  # 请求技术页面
        technique = get_technique(bs_technique)
        row_num = 1
        for h1 in technique:
            bs_sub_technique = request_url(h1["href"])  # 请求子技术页面
            sub_technique = get_sub_technique(bs_sub_technique)
            if sub_technique == None:
                ws.cell(row=row_num, column=1, value=h1["name"])  # 写入技术名称
                row_num += 1
            else:
                sub_technique_num = len(sub_technique)
                print(sub_technique)
                print(row_num)
                if sub_technique_num > 1:
                    ws.cell(row=row_num, column=1, value=h1["name"])  # 写入技术名称
                    for x, y in enumerate(sub_technique):
                        ws.cell(row=x + row_num, column=2, value=y)  # 写入子技术名称
                    row_num += sub_technique_num
                else:
                    ws.cell(row=row_num, column=1, value=h1["name"])  # 写入技术名称
                    for x, y in enumerate(sub_technique):
                        ws.cell(row=x + row_num, column=2, value=y)  # 写入子技术名称
                    row_num += 1
        wb.save('.\Spiderman\ATT&CK\AttckMatrix.xlsx')
        wb.close()
